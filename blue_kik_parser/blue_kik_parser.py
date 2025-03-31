#!/usr/bin/env python3

import sqlite3
import sys
import argparse
import logging
from datetime import datetime, timezone
import os
import io
import pandas as pd
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Alignment

__dtfmt__ = "%Y-%m-%d %H:%M:%S"

# Configure logging with debug level
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)-8s - %(message)s",
    encoding="utf-8",
    datefmt=__dtfmt__,
)

STD_QUERIES = {
    "Private Messages": """
        SELECT m._id, m.bin_id, m.timestamp,
               CASE WHEN m.was_me = 1 THEN 'account owner' ELSE m.partner_jid END AS sender,
               COALESCE(m.body, '[No Text]') AS body,
               m.stat_msg, m.stat_user_jid,
               m.content_id, kc.content_name, ku.content_uri, ai.image_id,
               m.friend_attr_id, m.was_me
        FROM messagesTable m
        LEFT JOIN KIKContentTable kc ON m.content_id = kc.content_id
        LEFT JOIN KIKContentURITable ku ON m.content_id = ku.content_id
        LEFT JOIN AccountSwitcherImgBackupTable ai ON kc.content_string = ai.image_id
        WHERE m.bin_id LIKE '%@talk.kik.com'
        AND (kc.content_name = 'preview' OR kc.content_name IS NULL);
    """,
    "Group Messages": """
        SELECT m._id, m.bin_id, m.timestamp,
               CASE WHEN m.was_me = 1 THEN 'account owner' ELSE m.partner_jid END AS sender,
               COALESCE(m.body, '[No Text]') AS body,
               m.stat_msg, m.stat_user_jid,
               m.content_id, kc.content_name, ku.content_uri, ai.image_id,
               m.friend_attr_id, m.was_me
        FROM messagesTable m
        LEFT JOIN KIKContentTable kc ON m.content_id = kc.content_id
        LEFT JOIN KIKContentURITable ku ON m.content_id = ku.content_id
        LEFT JOIN AccountSwitcherImgBackupTable ai ON kc.content_string = ai.image_id
        WHERE m.bin_id LIKE '%@groups.kik.com'
        AND (kc.content_name = 'preview' OR kc.content_name IS NULL);
    """,
    "Images": """
        SELECT CASE WHEN m.was_me = 1 THEN 'account owner' ELSE m.partner_jid END AS sender,
               m.content_id, kc.content_name, ku.content_uri, ai.image_id,
               kr.retain_count, m.was_me
        FROM messagesTable m
        LEFT JOIN KIKContentTable kc ON m.content_id = kc.content_id
        LEFT JOIN KIKContentURITable ku ON m.content_id = ku.content_id
        LEFT JOIN AccountSwitcherImgBackupTable ai ON kc.content_string = ai.image_id
        LEFT JOIN KIKContentRetainCountTable kr ON m.content_id = kr.content_id
        WHERE m.bin_id LIKE '%@talk.kik.com'
        AND kc.content_name NOT IN (
            'icon', 'app-name', 'file-name', 'file-size', 'int-file-url-local',
            'int-file-state', 'int-chunk-progress', 'file-url', 'sha1-scaled',
            'blockhash-scaled', 'sha1-original', 'allow-forward'
        );
    """,
}

HTML_QUERIES = {
    "Private Messages": """
        SELECT m._id, m.bin_id, DATETIME(m.timestamp / 1000, 'unixepoch') AS timestamp,
               CASE WHEN m.was_me = 1 THEN 'account owner' ELSE m.partner_jid END AS sender,
               COALESCE(m.body, '[No Text]') AS body,
               m.content_id, COALESCE(ai.image_id, kc.content_string) AS image_id
        FROM messagesTable m
        LEFT JOIN AccountSwitcherImgBackupTable ai ON m.content_id = ai.image_id
        LEFT JOIN KIKContentTable kc ON m.content_id = kc.content_id  -- Corrected join
        WHERE m.bin_id LIKE '%@talk.kik.com';
    """,
    "Group Messages": """
        SELECT m._id, m.bin_id, m.timestamp,
               CASE WHEN m.was_me = 1 THEN 'account owner' ELSE m.partner_jid END AS sender,
               COALESCE(m.body, '[No Text]') AS body,
               m.content_id, ai.image_id
        FROM messagesTable m
        LEFT JOIN AccountSwitcherImgBackupTable ai ON m.content_id = ai.image_id
        WHERE m.bin_id LIKE '%@groups.kik.com';
    """,
    "Images": """
        SELECT CASE WHEN m.was_me = 1 THEN 'account owner' ELSE m.partner_jid END AS sender,
               m.content_id, ai.image_id
        FROM messagesTable m
        LEFT JOIN AccountSwitcherImgBackupTable ai ON m.content_id = ai.image_id;
    """,
}

HTML_CONTENT = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Kik Backup Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        .tab { display: none; }
        .tab.active { display: block; }
        .tabs { display: flex; cursor: pointer; }
        .tab-button { padding: 10px; border: 1px solid #ddd; background: #f0f0f0; margin-right: 5px; }
        table { width: 100%; border-collapse: collapse; margin-bottom: 20px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f4f4f4; }
        img { max-width: 150px; height: auto; display: block; }
    </style>
    <script>
        function showTab(tabId) {
            document.querySelectorAll('.tab').forEach(tab => tab.classList.remove('active'));
            document.getElementById(tabId).classList.add('active');
        }
    </script>
</head>
<body>
    <h1>Kik Backup Report</h1>
    <div class="tabs">
        <div class="tab-button" onclick="showTab('private_messages')">Private Messages</div>
        <div class="tab-button" onclick="showTab('group_messages')">Group Messages</div>
        <div class="tab-button" onclick="showTab('images')">Images</div>
    </div>
"""


# Function to convert UNIX milliseconds timestamp to human-readable format
def convert_to_human_readable(unix_timestamp_ms):
    try:
        unix_timestamp = unix_timestamp_ms / 1000.0
        return datetime.fromtimestamp(unix_timestamp, tz=timezone.utc).strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    except (OSError, ValueError) as e:
        logging.error(f"Error converting timestamp {unix_timestamp_ms}: {e}")
        return None


# Function to check if a column exists in a table
def column_exists(table, column, cursor):
    cursor.execute(f"PRAGMA table_info({table})")
    columns = [row[1] for row in cursor.fetchall()]
    logging.debug(f"Columns in {table}: {columns}")
    return column in columns


def extract_images(database, cursor):
    output_dir = f"{os.path.splitext(os.path.basename(database))[0]}_images"
    os.makedirs(output_dir, exist_ok=True)

    # Execute the SQL query to fetch images
    cursor.execute("SELECT image_id, image_bytes FROM AccountSwitcherImgBackupTable")

    # Process images in batches to optimize memory usage
    batch_size = 100
    while True:
        count = 0
        rows = cursor.fetchmany(batch_size)
        if not rows:
            break
        for image_id, image_bytes in rows:
            if image_bytes:  # Ensure the image is not empty or None
                try:
                    # Determine file extension
                    file = Image.open(io.BytesIO(image_bytes))
                    file_extension = file.format.lower() or "jpg"
                    filename = os.path.join(output_dir, f"{image_id}.{file_extension}")
                    # Write image data to file
                    with open(filename, "wb") as file:
                        file.write(image_bytes)
                    logging.info(f"Saved: {filename}")
                    count += 1
                except Exception as e:
                    logging.error(f"Error saving image {image_id}: {e}")
            else:
                logging.warning(f"Skipping {image_id}: No image data")

    # Close the cursor and connection
    logging.info(f"Image extraction completed. {count} images extracted.")


def fetch_data_from_db(db_path, queries):
    """Fetches data from the SQLite database and returns a dictionary of DataFrames."""
    try:
        conn = sqlite3.connect(db_path)
        data_frames = {
            key: pd.read_sql_query(query, conn) for key, query in queries.items()
        }
        conn.close()
        logging.info("Successfully fetched data from database.")
        return data_frames
    except Exception as e:
        logging.error(f"Error fetching data from database: {e}")
        raise RuntimeError("Critical database error. Execution halted.") from e


def generate_html(data_frames, output_file, image_folder):
    """Generates an HTML report with tab navigation and saves it to a file."""
    content = HTML_CONTENT
    try:
        for section, df in data_frames.items():
            section_id = section.lower().replace(" ", "_")
            content += (
                f"<div id='{section_id}' class='tab'><h2>{section}</h2><table><tr>"
            )
            content += "".join(f"<th>{col}</th>" for col in df.columns)
            content += "</tr>"

            for _, row in df.iterrows():
                content += "<tr>"
                for col in df.columns:
                    if col == "image_id" and pd.notna(row[col]):
                        image_extensions = ["jpg", "jpeg", "png", "gif"]
                        image_path = None
                        for ext in image_extensions:
                            potential_path = os.path.join(
                                image_folder, f"{row[col]}.{ext}"
                            )
                            if os.path.exists(potential_path):
                                image_path = potential_path
                                break
                        if not image_path:
                            image_path = "Image Not Found"
                            logging.warning(f"Missing image: {row[col]}")
                        else:
                            logging.info(f"Image found: {image_path}")
                        content += f"<td><img src='{image_path}' alt='Image'></td>"
                content += "</tr>"
            content += "</table></div>"

        content += "<script>showTab('private_messages');</script></body></html>"

        with open(output_file, "w", encoding="utf-8") as f:
            f.write(content)

        logging.info(f"HTML report successfully generated: {output_file}")
    except Exception as e:
        logging.error(f"Error generating HTML report: {e}")
        raise


def main():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    parser = argparse.ArgumentParser(
        description="Extract and export messages from an SQLite database."
    )
    parser.add_argument("database", type=str, help="Path to the SQLite database file")
    parser.add_argument(
        "-e",
        "--extract",
        help="Extract images from the SQLite database",
        action="store_true",
        required=False,
    )
    parser.add_argument(
        "-i",
        "--images",
        help="Path to the folder containing the extracted images, default is 'kikdatabase_images' in the current path",
        default="kikdatabase_images",
        required=False,
    )
    parser.add_argument(
        "--html",
        help="Generate an HTML report from a Kik backup database - requires --images",
        action="store_true",
    )
    args = parser.parse_args()
    if len(sys.argv[1:]) == 0:
        parser.print_help()
        sys.exit(0)
    if not os.path.exists(args.database):
        logging.error(
            f"The database {args.database} does not exist. Please check your path and try again."
        )
        sys.exit(1)
    logging.debug(f"Connecting to database: {args.database}")
    conn = sqlite3.connect(args.database)
    cursor = conn.cursor()
    if not column_exists("messagesTable", "bin_id", cursor):
        logging.error("Required column 'bin_id' not found in messagesTable. Exiting.")
        sys.exit(1)
    output_file = os.path.join(
        os.path.dirname(args.database),
        f"Blue Kik Parsed - {os.path.basename(args.database)} - {timestamp}.xlsx",
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, query in STD_QUERIES.items():
            logging.info(f"Processing {sheet_name} details ...")
            logging.debug(f"Executing query for {sheet_name}: {query}")
            try:
                chunk_iter = pd.read_sql_query(query, conn, chunksize=1000)
                df_list = []
                for chunk in chunk_iter:
                    logging.debug(f"Retrieved {len(chunk)} rows for {sheet_name}")
                    if "timestamp" in chunk.columns:
                        chunk["timestamp"] = chunk["timestamp"].apply(
                            convert_to_human_readable
                        )
                    df_list.append(chunk)
                full_df = (
                    pd.concat(df_list, ignore_index=True) if df_list else pd.DataFrame()
                )
                full_df.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception as e:
                logging.error(f"Error processing {sheet_name}: {e}")
    wb = load_workbook(output_file)
    column_widths = {
        "_id": 10,
        "bin_id": 25,
        "timestamp": 20,
        "sender": 30,
        "body": 100,
        "stat_msg": 20,
        "stat_user_jid": 30,
        "content_id": 20,
        "content_name": 30,
        "content_uri": 50,
        "image_id": 20,
        "friend_attr_id": 20,
        "was_me": 10,
        "retain_count": 15,
    }

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.iter_cols(1, ws.max_column):
            col_letter = col[0].column_letter
            col_name = col[0].value
            if col_name in column_widths:
                ws.column_dimensions[col_letter].width = column_widths[col_name]
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Enable autofilter for all columns
        ws.auto_filter.ref = ws.dimensions
        wb.save(output_file)
    if args.extract:
        extract_images(args.database, cursor)
    if args.html:
        if not args.images:
            logging.error(
                "The -i / --images argument is required to generate an HTML report. If images haven't been extracted yet, use the -e / --extract argument."
            )
            sys.exit(1)
        if not os.path.exists(args.images) or not os.path.isdir(args.images):
            logging.error(
                f" The path {args.images} does not exist. Check your path and try again. If you haven't extracted images yet, use the -e / --extract argument."
            )
            sys.exit(1)
        OUTPUT_HTML = os.path.join(
            os.path.dirname(args.database),
            os.path.splitext(os.path.basename(args.database))[0] + "_Report.html",
        )
        data_frames = fetch_data_from_db(args.database, HTML_QUERIES)
        if data_frames:
            generate_html(data_frames, OUTPUT_HTML, args.images)
        else:
            logging.warning("No data extracted from the database.")
        logging.info("Script execution finished")
    # Close the cursor and connection
    logging.debug("Closing database connection.")
    cursor.close()
    conn.close()
    logging.info(f"Enhanced data export completed: '{os.path.normpath(output_file)}'")


if __name__ == "__main__":
    main()
