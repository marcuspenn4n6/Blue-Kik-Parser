import sqlite3
import pandas as pd
import argparse
import logging
from datetime import datetime, timezone
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Configure logging with debug level
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Function to convert UNIX milliseconds timestamp to human-readable format
def convert_to_human_readable(unix_timestamp_ms):
    try:
        unix_timestamp = unix_timestamp_ms / 1000.0
        return datetime.fromtimestamp(unix_timestamp, tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
    except (OSError, ValueError) as e:
        logging.error(f"Error converting timestamp {unix_timestamp_ms}: {e}")
        return None

# Parse command-line arguments
parser = argparse.ArgumentParser(description='Extract and export messages from an SQLite database.')
parser.add_argument('database', type=str, help='Path to the SQLite database file')
args = parser.parse_args()

# Generate output filename based on database argument
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = os.path.join(os.path.dirname(args.database),
                           f"Blue Kik Parsed - {os.path.basename(args.database)} - {timestamp}.xlsx")

# Connect to the SQLite database
logging.debug(f"Connecting to database: {args.database}")
conn = sqlite3.connect(args.database)
cursor = conn.cursor()

# Function to check if a column exists in a table
def column_exists(table, column):
    cursor.execute(f"PRAGMA table_info({table})")
    columns = [row[1] for row in cursor.fetchall()]
    logging.debug(f"Columns in {table}: {columns}")
    return column in columns

# Ensure bin_id exists in messagesTable
bin_id_column = "bin_id" if column_exists("messagesTable", "bin_id") else None
if not bin_id_column:
    logging.error("Required column 'bin_id' not found in messagesTable. Exiting.")
    exit(1)

queries = {
    "Private Messages": """
        SELECT m._id, m.bin_id, m.timestamp, 
               CASE WHEN m.was_me = 1 THEN 'account owner' ELSE m.partner_jid END AS sender, 
               COALESCE(m.body, '[No Text]') AS body, 
               m.stat_msg, m.stat_user_jid, 
               m.content_id, kc.content_name, ku.content_uri, ai.image_id, 
               m.friend_attr_id, m.was_me
        FROM messagesTable m
        LEFT JOIN KIKContentTable kc ON m.content_id = kc.content_id
        LEFT JOIN KIKContentURITable ku ON m.content_id = ku.content_id
        LEFT JOIN AccountSwitcherImgBackupTable ai ON kc.content_string = ai.image_id
        WHERE m.bin_id LIKE '%@talk.kik.com' 
        AND (kc.content_name = 'preview' OR kc.content_name IS NULL);
    """,

    "Group Messages": """
        SELECT m._id, m.bin_id, m.timestamp, 
               CASE WHEN m.was_me = 1 THEN 'account owner' ELSE m.partner_jid END AS sender, 
               COALESCE(m.body, '[No Text]') AS body, 
               m.stat_msg, m.stat_user_jid, 
               m.content_id, kc.content_name, ku.content_uri, ai.image_id, 
               m.friend_attr_id, m.was_me
        FROM messagesTable m
        LEFT JOIN KIKContentTable kc ON m.content_id = kc.content_id
        LEFT JOIN KIKContentURITable ku ON m.content_id = ku.content_id
        LEFT JOIN AccountSwitcherImgBackupTable ai ON kc.content_string = ai.image_id
        WHERE m.bin_id LIKE '%@groups.kik.com' 
        AND (kc.content_name = 'preview' OR kc.content_name IS NULL);
    """,

    "Images": """
        SELECT CASE WHEN m.was_me = 1 THEN 'account owner' ELSE m.partner_jid END AS sender, 
               m.content_id, kc.content_name, ku.content_uri, ai.image_id, 
               kr.retain_count, m.was_me
        FROM messagesTable m
        LEFT JOIN KIKContentTable kc ON m.content_id = kc.content_id
        LEFT JOIN KIKContentURITable ku ON m.content_id = ku.content_id
        LEFT JOIN AccountSwitcherImgBackupTable ai ON kc.content_string = ai.image_id
        LEFT JOIN KIKContentRetainCountTable kr ON m.content_id = kr.content_id
        WHERE m.bin_id LIKE '%@talk.kik.com' 
        AND kc.content_name NOT IN (
            'icon', 'app-name', 'file-name', 'file-size', 'int-file-url-local', 
            'int-file-state', 'int-chunk-progress', 'file-url', 'sha1-scaled', 
            'blockhash-scaled', 'sha1-original', 'allow-forward'
        );
    """
}

# Process and export queries in chunks
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name, query in queries.items():
        logging.info(f"Processing {sheet_name}...")
        logging.debug(f"Executing query for {sheet_name}: {query}")
        try:
            chunk_iter = pd.read_sql_query(query, conn, chunksize=1000)
            df_list = []
            for chunk in chunk_iter:
                logging.debug(f"Fetched {len(chunk)} rows for {sheet_name}")
                if 'timestamp' in chunk.columns:
                    chunk['timestamp'] = chunk['timestamp'].apply(convert_to_human_readable)
                df_list.append(chunk)
            full_df = pd.concat(df_list, ignore_index=True) if df_list else pd.DataFrame()
            full_df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            logging.error(f"Error processing {sheet_name}: {e}")

# Close the cursor and connection
logging.debug("Closing database connection.")
cursor.close()
conn.close()

# Apply column width formatting
wb = load_workbook(output_file)
column_widths = {
    '_id': 10, 'bin_id': 25, 'timestamp': 20, 'sender': 30, 'body': 100,
    'stat_msg': 20, 'stat_user_jid': 30, 'content_id': 20, 'content_name': 30,
    'content_uri': 50, 'image_id': 20, 'friend_attr_id': 20, 'was_me': 10,
    'retain_count': 15
}

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in ws.iter_cols(1, ws.max_column):
        col_letter = col[0].column_letter
        col_name = col[0].value
        if col_name in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[col_name]
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    # Enable autofilter for all columns
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_file)

logging.info(f"Enhanced data export completed: {output_file}")